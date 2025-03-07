import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font
from excel_to_pdf import convert_to_pdf
import os

def abs_path(f_name):
    return os.path.abspath(f_name) 

df = pd.read_excel("test.xlsx")

# creating a new dataframe df1 and copying the contents from text.xlsx to df1 through df
df1 = pd.DataFrame()

#df1["q_no"]  = df["Q_No"].str.extract(r'(\d+)')
df1["q_no"] = df["Q_No"].apply(lambda x: f"Q.{x[:-1]}" if x[-1] == 'a' else " ")
df1["sub_q"] = df["Q_No"].apply(lambda x: f"{x[-1]}.")   

df1["Question"] = df["Question"].copy()
df1["Marks"] = df["Marks"].copy()
df1["L"] = df["L"].copy()
df1["CO"] = df["CO"].copy()

df1["L"] = df["L"].str.upper()
df1["CO"] = df["CO"].str.upper()

# saving the dataframe into excel sheet
file_name = "output_qp_format.xlsx"
df1.to_excel(file_name, index=False, header=False, sheet_name="Sheet1")

# loading Worksheet into python
wb = load_workbook(file_name)
ws = wb.active

# Fixing the width of each column
ws.column_dimensions["A"].width = 6  
ws.column_dimensions["B"].width = 3  
ws.column_dimensions["C"].width = 75  
ws.column_dimensions["D"].width = 5
ws.column_dimensions["E"].width = 5
ws.column_dimensions["F"].width = 6

border_style = Border(left=Side(style="thin"), 
                      right=Side(style="thin"),
                      top=Side(style="thin"), 
                      bottom=Side(style="thin"))

# To get the row position where MODULE/OR value is to be added
pos = []
for i in range(1, ws.max_row+1):
    if ws[f"B{i}"].value == "a.":
        pos.append(i)

# for merging cells to add MODULE and OR cell values
step = 0
mod_cnt = 2
for i in pos:
    row_pos = i + step
    ws.insert_rows(row_pos)
    ws.merge_cells(f"A{row_pos}:F{row_pos}")
    ws[f"A{row_pos}"].value = "MODULE - 1"
    
    if row_pos>1:
        cell_val = str(ws[f"A{row_pos+1}"].value)
        numbers = re.findall(r'\d+', cell_val)  
        int_part = int("".join(numbers)) if numbers else None  
        
        if int_part % 2 == 0:
            ws[f"A{row_pos}"].value = f"MODULE - {mod_cnt}"
            mod_cnt += 1
        else:
            ws[f"A{row_pos}"].value = "OR"
    
    for row in ws[f"A{row_pos}:F{row_pos}"]:
        for cell in row:
            cell.border = border_style 
            
    ws[f"A{row_pos}"].alignment = Alignment(horizontal="center", vertical="center")
        
    step += 1

# Alignment
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column): 
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=3, max_col=3): 
    for cell in row:
        cell.alignment = Alignment(horizontal="left", vertical="center")

# Text wrapping for lengthy questions
for col in [3, 5]:
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col, max_col=col):
        for cell in row:
            if col == 5:
                cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(wrap_text=True, vertical="center")
     
# Change Font Style to Times New Roman   
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        if cell.value:  
            cell.font = Font(name="Times New Roman", bold = True)
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=3, max_col=3): 
    for cell in row:
        cell.font = Font(name="Times New Roman")
        
# Adding Border
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        if cell.value != None:
            cell.border = border_style

# Saving the formats made above
wb.save(file_name)
print(f"{file_name} created successfully!")

convert_to_pdf(abs_path(file_name))

