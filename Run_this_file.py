import pandas as pd
import os

from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.drawing.image import Image

from excel_to_pdf import convert_to_pdf     # it is just another python file which contains the function to convert excel sheet to pdf

################################################################################################################################################################################################################

# below code section's function is to:
# 1) collect the content from the test.xlsx file 
# 2) load it into a dataframe 
# 3) rearrange it into a proper format 
# 4) Store other details like sem, subject, time, etc. into variables
# 5) Store the rearranged content into a new excel file - "output_qp_format.xlsx" 

def abs_path(f_name):                           # required because the win32com module int the excel_to_pdf.py file takes in only absolute paths
    return os.path.abspath(f_name) 

df = pd.read_excel("input_excel_file.xlsx") # Please enter the name or path of the input excel file. 

# creating a new dataframe - df1, and copying the contents from input_excel_file.xlsx to df1 through df
df1 = pd.DataFrame()

#df1["q_no"]  = df["Q_No"].str.extract(r'(\d+)')
df1["q_no"] = df["Q_No"].apply(lambda x: int(f"{x[:-1]}") if x[-1].lower() == 'a' else " ") # selects only the integer (1,2,3...) part and stores it in q_no series
df1["sub_q"] = df["Q_No"].apply(lambda x: f"{x[-1].lower()}.")   # selects the alphabet (a/b/c...) part and stores it in sub_q series

# using .copy() for deepcopy just for safety
df1["Question"] = df["Question"].copy()
df1["Marks"] = df["Marks"].copy()
df1["L"] = df["L"].copy()
df1["CO"] = df["CO"].copy()

df1["L"] = df["L"].str.upper()
df1["CO"] = df["CO"].str.upper()

# storing other details into their respective variables for their future use
semester = df["Semester"].iloc[0]  
subject = df["Subject"].iloc[0]
subject_code = df["Subject_code"].iloc[0]
exam_date = df["Date"].iloc[0]
exam_time = df["Time"].iloc[0]
department = df["Department"].iloc[0]
test_no = df["Test_no"].iloc[0]

# saving the dataframe into excel sheet
file_name = "output_qp_format.xlsx"
df1.to_excel(file_name, index=False, header=False, sheet_name="Sheet1")

################################################################################################################################################################################################################

# below codes's function is to:
# 1) Fixing the width of each column
# 2) Add 2 images rnsit_logo.jpeg and usn.jpeg
# 3) Add RNSIT Header (About RNSIT, Address, Ph. No.)
# 4) Display the stored variables of sem, subject, time, etc. 
# 5) Add new merged cell with value "MODULE - no." or "OR" after every two question numbers
# 6) Basic text formating like:
#             - Text alignment 
#             - converting to Bold
#             - changing font style to "Times New Roman"
#             - Text wrapping for the "Question" column 
# 7) Adding border to the whole content
# 8) Saving the excel file and converting it to PDF by calling a function from excel_to_pdf.py file

# loading Worksheet into python
wb = load_workbook(file_name)
ws = wb.active

# Fixing the width of each column
ws.column_dimensions["A"].width = 4.5 
ws.column_dimensions["B"].width = 3  
ws.column_dimensions["C"].width = 75  
ws.column_dimensions["D"].width = 6
ws.column_dimensions["E"].width = 6
ws.column_dimensions["F"].width = 6

border_style = Border(left=Side(style="thin"), 
                      right=Side(style="thin"),
                      top=Side(style="thin"), 
                      bottom=Side(style="thin"))

# Institution Details
RNSIT_header = [
    "RN SHETTY TRUST\u00AE", # \u00AE is the unicode for the encircled "R" trade mark
    "RNS INSTITUTE OF TECHNOLOGY",
    "Autonomous Institution Affiliated to VTU, Recognized by GOK, Approved by AICTE",
    "(NAAC 'A+ Grade' Accredited, NBA Accredited (UGCSE, ECE, ISE, EIE and EEE))",
    "Channasandra, Dr. Vishnuvardhan Road, Bengaluru 560 098",
    "Ph:(080)28611880,28611881 URL: www.rnsit.ac.in"
]

# Insert rows at the top for the institution details
ws.insert_rows(1, amount=17)

# Adding the content
for row, text in enumerate(RNSIT_header, start=1):
    ws.merge_cells(f"B{row}:F{row}")
    ws[f"B{row}"] = text
    ws[f"B{row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws[f"B{row}"].font = Font(name="Times New Roman", bold=True if row <= 2 else False, size=13 if row <= 2 else 11)
for i in range(1, 7):
    ws.row_dimensions[i].height = 12.5

# RNSIT logo image insertion
img = Image("rnsit_logo.jpeg")   
img.width = 110
img.height = 100
ws.add_image(img, "A1") 

# USN image insertion
img2 = Image("usn.jpeg")  
img2.width = 280
img2.height = 30
ws.add_image(img2, "C11")

ws.row_dimensions[7].height = 5

# Merge A8 to F8 and add "Department"
ws.merge_cells("A8:F8")
ws["A8"] = "Department of " + department
ws["A8"].alignment = Alignment(horizontal="center", vertical="center")
ws["A8"].font = Font(name="Times New Roman", bold=True, size=16)
ws.row_dimensions[8].height = 23

# code for the brown line
for row in ws.iter_rows(min_row=9, max_row=9, min_col=2, max_col=5):
    for cell in row:
        cell.fill = PatternFill(start_color="A52A2A", end_color="A52A2A", fill_type="solid")
ws.row_dimensions[9].height = 3
ws.row_dimensions[10].height = 3

# displaying stored variables
ws.merge_cells("A13:F13")
ws["A13"] = "CIE - Test " + str(int(test_no))
ws["A13"].alignment = Alignment(horizontal="center", vertical="center")
ws["A13"].font = Font(name="Times New Roman", bold=True, size=14)

ws["A14"] = "Semester: " + semester
ws["A15"] = "Subject: " + subject
ws["A16"] = "Subject Code: " + subject_code
ws["D14"] = "Date: " + exam_date
ws["D15"] = "Time: " + exam_time  
ws["D16"] = "Max. Marks: 50 "
for i in range(4,7):
    # Error # ws.row_dimensions[int("1"+str(i))] = 11
    ws["A1"+str(i)].font = Font(name="Times New Roman", bold=True)
    ws["D1"+str(i)].font = Font(name="Times New Roman", bold=True)

# adding an instruction line 
ws["A17"] = "Instruction: Answer any 5 full questions, selecting ONE question from each part."
ws["A17"].font = Font(name="Times New Roman", bold=True, italic=True)
# Error # ws.row_dimensions[17] = 11

# To get the row position where "OR" value is to be added
pos = []
for i in range(1, ws.max_row+1):
    if ws[f"B{i}"].value == "a.":
        pos.append(i)

# for merging cells to add "OR" cell value
step = 0
mod_cnt = 2
for i in pos:
    row_pos = i + step
    
    if step == 0:
        ws.insert_rows(row_pos)
        ws.row_dimensions[row_pos].height = 19
        ws.merge_cells(f"A{row_pos}:B{row_pos}")
        ws[f"A{row_pos}"] = "Q. No."
        ws[f"C{row_pos}"] = "Questions"
        ws[f"D{row_pos}"] = "Marks"
        ws[f"E{row_pos}"] = "RBT*"
        ws[f"F{row_pos}"] = "COs"
        for row in ws[f"A{row_pos}:F{row_pos}"]:
            for cell in row:
                cell.border = border_style 
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(name="Times New Roman", bold=True, size=10)
        step += 1
        continue

    ws.insert_rows(row_pos)
    ws.merge_cells(f"A{row_pos}:F{row_pos}")
    
    if ws[f"A{row_pos+1}"].value % 2 == 0:     # This part exclusively adds "OR" between the specific two question numbers like 1-2, 3-4, etc.
            ws[f"A{row_pos}"] = "OR"
    
    for row in ws[f"A{row_pos}:F{row_pos}"]:
        for cell in row:
            cell.border = border_style 
            
    ws[f"A{row_pos}"].alignment = Alignment(horizontal="center", vertical="center")
        
    step += 1
    
# from here the min_row is considered to be 19 because the actual table containing questions start from row 19

# Alignment
for row in ws.iter_rows(min_row=19, max_row=ws.max_row, min_col=1, max_col=ws.max_column): 
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")
for row in ws.iter_rows(min_row=19, max_row=ws.max_row, min_col=3, max_col=3): 
    for cell in row:
        cell.alignment = Alignment(horizontal="left", vertical="center")

# Text wrapping for lengthy questions
for col in [3, 5]:
    for row in ws.iter_rows(min_row=19, max_row=ws.max_row, min_col=col, max_col=col):
        for cell in row:
            if col == 5:
                cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(wrap_text=True, vertical="center")
     
# Change Font Style to Times New Roman   
for row in ws.iter_rows(min_row=19, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        if cell.value:  
            cell.font = Font(name="Times New Roman", bold = True)
for row in ws.iter_rows(min_row=19, max_row=ws.max_row, min_col=3, max_col=3): 
    for cell in row:
        cell.font = Font(name="Times New Roman")
        
# Adding Border
for row in ws.iter_rows(min_row=19, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        if cell.value != None:
            cell.border = border_style
          
# adding the information regarding RBT at the last row
max_r = ws.max_row + 1           
ws.merge_cells(f"A{max_r}:F{max_r}")            
ws[f"A{max_r}"] = "*Revised Bloom’s Taxonomy: L1-Remember, L2-Understand, L3-Apply, L4-Analyze, L5-Evaluate, L6-Create"
ws[f"A{max_r}"].alignment = Alignment(horizontal="center", vertical="center")
ws[f"A{max_r}"].font = Font(name="Times New Roman", bold=True, size=8)

# Saving the formats made above
wb.save(file_name)
print(f"{file_name} created successfully!")

# function call to excel_to_pdf.py 
convert_to_pdf(abs_path(file_name))

